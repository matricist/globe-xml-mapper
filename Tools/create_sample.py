"""
테스트용 샘플 데이터 생성 스크립트.
디렉토리 구조:
  sample/
  ├── sample_기본정보.xlsx     (루트, 시트 "1.1~1.2")
  ├── 1.3.1/
  │   ├── upe_001.xlsx
  │   └── upe_002.xlsx
  ├── 1.3.2.1/
  │   ├── ce_001.xlsx
  │   ├── ce_002.xlsx
  │   └── ce_003.xlsx
  └── 1.3.2.2/
      └── ex_001.xlsx
"""

import os
import shutil
from openpyxl import load_workbook

BASE_DIR = os.path.join(os.path.dirname(__file__), "..")
TEMPLATES_DIR = os.path.join(BASE_DIR, "Resources", "templates")
SAMPLE_DIR = os.path.join(BASE_DIR, "Resources", "sample")


def copy_template(template_name, dest_path):
    src = os.path.join(TEMPLATES_DIR, template_name)
    shutil.copy2(src, dest_path)
    return dest_path


def fill_sheet(path, sheet_name, data):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    for cell, val in data.items():
        ws[cell] = val
    wb.save(path)


def fill_attachment(path, main_sheet, exclude_sheets, owners):
    """별첨 시트에 소유지분 데이터 기재"""
    wb = load_workbook(path)
    attach = None
    for s in wb.worksheets:
        if s.title not in exclude_sheets:
            attach = s
            break
    if attach is None:
        attach = wb.create_sheet("별첨")
    for i, (owner_type, tin, pct) in enumerate(owners, 3):
        attach.cell(row=i, column=2, value=owner_type)
        attach.cell(row=i, column=3, value=tin)
        attach.cell(row=i, column=4, value=pct)
    wb.save(path)


def clean_sample_dir():
    if os.path.exists(SAMPLE_DIR):
        shutil.rmtree(SAMPLE_DIR)
    os.makedirs(SAMPLE_DIR)
    for sub in ["1.3.1", "1.3.2.1", "1.3.2.2"]:
        os.makedirs(os.path.join(SAMPLE_DIR, sub))


def main():
    clean_sample_dir()
    print("샘플 데이터 생성 중...")

    # === 루트: 기본정보 ===
    path = copy_template("template_1.1~1.2.xlsx", os.path.join(SAMPLE_DIR, "기본정보.xlsx"))
    fill_sheet(path, "1.1~1.2", {
        "C3": "2025-01-01", "C6": "2025-12-31",
        "P3": "테스트주식회사", "P5": "123-45-67890",
        "B10": "GIR401", "D10": "Test Corp", "H10": "1234567890",
        "K10": "GIR401", "M10": "KR", "O10": "KR, JP, US",
        "B15": "테스트그룹", "F15": "2025-01-01", "K15": "2025-12-31", "O15": "GIR101",
        "B18": "GIR501", "H18": "K-IFRS", "M18": "KRW",
    })
    print(f"  생성: {path}")

    # === 1.3.1: UPE 2개 ===
    path = copy_template("template_1.3.1.xlsx", os.path.join(SAMPLE_DIR, "1.3.1", "upe_001.xlsx"))
    fill_sheet(path, "1.3.1", {
        "J22": "KR", "J23": "GIR201", "J24": "Test Ultimate Parent",
        "J25": "1234567890", "J26": "1234567890",
        "J27": "GIR301", "J28": "GIR601", "J29": "KR",
    })
    print(f"  생성: {path}")

    path = copy_template("template_1.3.1.xlsx", os.path.join(SAMPLE_DIR, "1.3.1", "upe_002.xlsx"))
    fill_sheet(path, "1.3.1", {
        "J22": "US", "J23": "GIR202", "J24": "Test Parent US",
        "J25": "US9876543210", "J26": "US9876543210", "J27": "GIR304",
    })
    print(f"  생성: {path}")

    # === 1.3.2.1: CE 3개 ===
    path = copy_template("template_1.3.2.1.xlsx", os.path.join(SAMPLE_DIR, "1.3.2.1", "ce_001.xlsx"))
    fill_sheet(path, "1.3.2.1", {
        "O5": "true", "O6": "JP", "O7": "GIR201",
        "O8": "Test Subsidiary Japan", "O9": "JP1234567890",
        "O10": "JP1234567890", "O11": "GIR301", "O19": "false",
    })
    fill_attachment(path, "1.3.2.1", {"1.3.2.1", "작성요령"},
        [("GIR802", "1234567890", 80), ("GIR801", "9876543210", 20)])
    print(f"  생성: {path}")

    path = copy_template("template_1.3.2.1.xlsx", os.path.join(SAMPLE_DIR, "1.3.2.1", "ce_002.xlsx"))
    fill_sheet(path, "1.3.2.1", {
        "O5": "false", "O6": "DE", "O7": "GIR204",
        "O8": "Test GmbH Germany", "O9": "DE5555555555",
        "O10": "DE5555555555", "O11": "GIR301",
    })
    fill_attachment(path, "1.3.2.1", {"1.3.2.1", "작성요령"},
        [("GIR802", "1234567890", 100)])
    print(f"  생성: {path}")

    path = copy_template("template_1.3.2.1.xlsx", os.path.join(SAMPLE_DIR, "1.3.2.1", "ce_003.xlsx"))
    fill_sheet(path, "1.3.2.1", {
        "O5": "false", "O6": "KR", "O7": "GIR201, GIR204",
        "O8": "Test Korea Sub", "O9": "KR7777777777",
        "O10": "KR7777777777", "O11": "GIR301",
        "O16": "GIR901", "O19": "true", "O20": "60", "O21": "true",
    })
    fill_attachment(path, "1.3.2.1", {"1.3.2.1", "작성요령"},
        [("GIR802", "1234567890", 60), ("GIR802", "JP1234567890", 40)])
    print(f"  생성: {path}")

    # === 1.3.2.2: 제외기업 1개 ===
    path = copy_template("template_1.3.2.2.xlsx", os.path.join(SAMPLE_DIR, "1.3.2.2", "ex_001.xlsx"))
    fill_sheet(path, "1.3.2.2", {
        "O4": "false",
        "O5": "Test Pension Fund",
        "O6": "GIR1004",
    })
    print(f"  생성: {path}")

    print("완료!")


if __name__ == "__main__":
    main()
