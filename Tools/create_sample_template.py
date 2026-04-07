"""
sample_template.xlsx 생성 — template.xlsx 복사 후 모든 시트에 샘플 데이터 채움.
+ 각 시트에 작성요령 시트 추가.
"""
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.join(os.path.dirname(__file__), "..")
TEMPLATE = os.path.join(BASE_DIR, "Resources", "template.xlsx")
OUTPUT = os.path.join(BASE_DIR, "Resources", "sample_template.xlsx")

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")


def fill_data(wb):
    """각 시트에 샘플 데이터 채움"""

    # === 1.1~1.2 (시트1) ===
    ws = wb["1.1~1.2"]
    data = {
        "C3": "2025-01-01", "C6": "2025-12-31",
        "P3": "테스트주식회사", "P5": "123-45-67890",
        "B10": "GIR401", "D10": "Test Corp", "H10": "1234567890",
        "K10": "GIR401", "M10": "KR", "O10": "KR, JP, US",
        "B15": "테스트그룹", "F15": "2025-01-01", "K15": "2025-12-31",
        "O15": "GIR101",
        "B18": "GIR501", "H18": "K-IFRS", "M18": "KRW",
    }
    for cell, val in data.items():
        ws[cell] = val

    # === 1.3.1 (시트2 — UPE, 데이터셀 J열 병합 J:R) ===
    ws = wb["1.3.1"]
    ws["J4"] = "KR"
    ws["J5"] = "GIR201"
    ws["J6"] = "Test Ultimate Parent"
    ws["J7"] = "1234567890"
    ws["J8"] = "1234567890"
    ws["J9"] = "GIR301"
    ws["J10"] = "GIR601"
    ws["J11"] = "KR"

    # === 1.3.2.1 (시트3 — CE 3개, 행 블록은 프로그램에서 추가해야 하므로 CE1만 직접 입력) ===
    # CE1: 일본 자회사
    ws = wb["1.3.2.1"]
    ws["O5"] = "TRUE"
    ws["O6"] = "JP"
    ws["O7"] = "GIR201"
    ws["O8"] = "Test Subsidiary Japan"
    ws["O9"] = "JP1234567890"
    ws["O10"] = "JP1234567890"
    ws["O11"] = "GIR301"
    ws["O14"] = "첨부1"
    ws["O19"] = "FALSE"

    # === 1.3.2.1 첨부 — 첨부1: 주주 3명 ===
    attach = None
    for s in wb.worksheets:
        if "첨부" in s.title:
            attach = s
            break
    if attach:
        # 주주 1
        attach["B5"] = "GIR802"
        attach["C5"] = "1234567890"
        attach["D5"] = "50"
        # 주주 2 (행 추가)
        attach["B6"] = "GIR801"
        attach["C6"] = "9876543210"
        attach["D6"] = "30"
        # 주주 3 (행 추가)
        attach["B7"] = "GIR802"
        attach["C7"] = "JP9999999999"
        attach["D7"] = "20"

    # === 1.3.2.2 (제외기업 — 데이터셀 O열) ===
    ws = wb["1.3.2.2"]
    ws["O4"] = "FALSE"
    ws["O5"] = "Test Pension Fund"
    ws["O6"] = "GIR1004"

    # === 1.3.3 (기업구조 변동 — 2행) ===
    ws = wb["1.3.3"]
    ws["P4"] = "TRUE"
    # 변동 1
    ws["B7"] = "Test Corp"
    ws["D7"] = "1234567890"
    ws["F7"] = "2025-06-01"
    ws["H7"] = "GIR301"
    ws["J7"] = "GIR302"
    ws["K7"] = "Parent Corp"
    ws["M7"] = "80"
    ws["P7"] = "60"
    # 변동 2
    ws["B8"] = "Test Sub JP"
    ws["D8"] = "JP1234567890"
    ws["F8"] = "2025-09-15"
    ws["H8"] = "GIR301"
    ws["J8"] = "GIR316"
    ws["K8"] = "Test Corp"
    ws["M8"] = "100"
    ws["P8"] = "0"

    # === 1.4 (정보 요약 — 3행) ===
    ws = wb["1.4"]
    # 국가 1: KR
    ws["B4"] = "KR"
    ws["I4"] = "GIR1201"
    ws["K4"] = "GIR1307"
    ws["M4"] = "TRUE"
    ws["O4"] = "GIR1401"
    ws["Q4"] = "GIR1501"
    # 국가 2: JP
    ws["B5"] = "JP"
    ws["K5"] = "GIR1306"
    ws["M5"] = "FALSE"
    ws["O5"] = "GIR1403"
    ws["Q5"] = "GIR1503"
    # 국가 3: DE
    ws["B6"] = "DE"
    ws["I6"] = "GIR1202"

    # === 시트 2 (국가별 적용면제) ===
    ws = wb["2"]
    # 2.1 기본사항
    ws["O6"] = "KR"
    ws["O9"] = "KR"
    ws["O10"] = "FALSE"
    # 2.2.1 적용면제
    ws["O15"] = "GIR1201"
    # 2.2.1.2 간소화 계산
    ws["H19"] = "5000000"
    ws["N19"] = "800000"
    ws["H20"] = "4800000"
    ws["N20"] = "750000"
    ws["H21"] = "4500000"
    ws["N21"] = "700000"
    ws["H22"] = "4766667"
    ws["N22"] = "750000"
    # 2.2.2 최소적용제외
    ws["E39"] = "5000000"
    ws["I39"] = "4900000"
    ws["L39"] = "1000000"
    ws["O39"] = "950000"


def _unused_add_guide_sheets(wb):
    """각 데이터 시트 뒤에 작성요령 정보를 별도 시트로 추가"""

    guides = {
        "1.1~1.2 작성요령": [
            ["헤더", "-", "사업연도 시작", "C3", "날짜 (YYYY-MM-DD)", "X", "-", "-"],
            ["헤더", "-", "사업연도 종료", "C6", "날짜 (YYYY-MM-DD)", "X", "-", "-"],
            ["헤더", "-", "법인명(상호)", "P3", "텍스트", "X", "-", "-"],
            ["헤더", "-", "사업자등록번호", "P5", "텍스트", "X", "-", "-"],
            ["1.1", "1", "최종모기업 여부", "B10", "Enum", "X", "GIR401=UPE, GIR402~405", "FilingCeRoleEnumType"],
            ["1.1", "2", "신고구성기업 상호", "D10", "텍스트", "X", "-", "-"],
            ["1.1", "3", "납세자번호", "H10", "텍스트", "X", "-", "-"],
            ["1.1", "4", "신고구성기업 지위", "K10", "Enum", "X", "GIR401~GIR405", "FilingCeRoleEnumType"],
            ["1.1", "5", "소재지국", "M10", "ISO 국가코드", "X", "KR, JP, US 등", "CountryCodeType"],
            ["1.1", "6", "교환당사국", "O10", "ISO 국가코드", "O (쉼표)", "KR, JP, US 등", "복수 시 쉼표 구분"],
            ["1.2.1", "1", "다국적기업그룹 명칭", "B15", "텍스트", "X", "-", "-"],
            ["1.2.1", "2", "사업연도 개시일", "F15", "날짜", "X", "-", "-"],
            ["1.2.1", "3", "사업연도 종료일", "K15", "날짜", "X", "-", "-"],
            ["1.2.1", "4", "수정신고서 여부", "O15", "Enum", "X", "GIR101=신규, GIR102=정정, GIR103=없음", "MessageTypeIndicEnumType"],
            ["1.2.2", "1", "연결재무제표 유형", "B18", "Enum", "X", "GIR501~504", "FilingCeCofUpeEnumType"],
            ["1.2.2", "2", "회계기준", "H18", "텍스트", "X", "-", "예: K-IFRS"],
            ["1.2.2", "3", "통화", "M18", "ISO 통화코드", "X", "KRW, USD 등", "CurrCodeType"],
        ],
        "1.3.1 작성요령": [
            ["", "", "", "", "", "", "", "Control Panel에서 [최종모기업 추가] 버튼으로 행 블록(3~11행) 추가 가능"],
            ["1.3.1", "1", "소재지국", "J3", "ISO 국가코드", "X", "KR, JP 등", "70010: 하나만"],
            ["1.3.1", "2", "적용가능규칙", "J4", "Enum", "O (쉼표)", "GIR201~205", "IdTypeRulesEnumType"],
            ["1.3.1", "3", "상호", "J5", "텍스트", "X", "-", "-"],
            ["1.3.1", "4", "납세자번호", "J6", "텍스트", "O (쉼표)", "-", "60022: FilingCE TIN 일치"],
            ["1.3.1", "5", "신고접수국가 TIN", "J7", "텍스트", "X", "-", "issuedBy=KR 자동"],
            ["1.3.1", "6", "기업유형", "J8", "Enum", "O (쉼표)", "GIR301~318", "70009: UPE 제한"],
            ["1.3.1", "7", "제외기업 유형", "J9", "Enum", "X", "GIR601~606", "ExcludedUpeEnumType"],
            ["1.3.1", "8", "시행령 적용", "J10", "ISO 국가코드", "X", "KR 등", "-"],
        ],
        "1.3.2.1 작성요령": [
            ["", "", "", "", "", "", "", "Control Panel에서 [구성기업 추가] 버튼으로 행 블록(4~21행) 추가. 첨부 시트에 첨부N 자동 생성"],
            ["변동", "1", "변동 여부", "O5", "bool", "X", "TRUE/FALSE", "-"],
            ["소재지국", "2", "소재지국", "O6", "ISO 국가코드", "X", "KR, JP 등", "70011: 하나만"],
            ["소재지국", "3", "적용가능규칙", "O7", "Enum", "O (쉼표)", "GIR201~205", "70012"],
            ["기본정보", "4", "구성기업 상호", "O8", "텍스트", "X", "-", "-"],
            ["기본정보", "5", "납세자번호", "O9", "텍스트", "X", "-", "-"],
            ["기본정보", "6", "신고접수국가 TIN", "O10", "텍스트", "X", "-", "issuedBy=KR 자동"],
            ["기본정보", "7", "기업유형", "O11", "Enum", "O (쉼표)", "GIR301~318", "70013~70021"],
            ["소유지분", "8~10", "소유지분", "첨부 시트", "별도", "O", "-", "첨부 시트에서 주주 행 추가. 퍼센트 입력"],
            ["QIIR", "11", "모기업 유형", "O16", "Enum", "X", "GIR901~903", "PopeipeEnumType"],
            ["QIIR", "12", "QIIR 모기업 TIN", "O17", "텍스트", "X", "-", "70033"],
            ["QUTPR", "14", "해외진출 초기 특례", "O19", "bool", "X", "TRUE/FALSE", "-"],
            ["QUTPR", "15", "소유지분 합계(%)", "O20", "숫자(%)", "X", "퍼센트 입력", "-"],
            ["QUTPR", "16", "UPE소유지분>합계", "O21", "bool", "X", "TRUE/FALSE", "-"],
        ],
        "1.3.2.2 작성요령": [
            ["", "", "", "", "", "", "", "Control Panel에서 [제외기업 추가] 버튼으로 행 블록(3~6행) 추가"],
            ["제외기업", "1", "변동 여부", "O3", "bool", "X", "TRUE/FALSE", "-"],
            ["제외기업", "2", "제외기업 상호", "O4", "텍스트", "X", "-", "-"],
            ["제외기업", "3", "제외기업 유형", "O5", "Enum", "X", "GIR1001~1008", "ExcludedEntityEnumType"],
        ],
        "1.3.3 작성요령": [
            ["", "", "", "", "", "", "", "Control Panel에서 [+] 버튼으로 데이터 행 추가 (7행부터)"],
            ["1.3.3", "1", "구성기업 상호", "B7", "텍스트", "X", "-", "-"],
            ["1.3.3", "2", "납세자번호", "D7", "텍스트", "X", "-", "-"],
            ["1.3.3", "3", "변동효력발생일", "F7", "날짜", "X", "-", "-"],
            ["1.3.3", "4", "변동 전 기업유형", "H7", "Enum", "X", "GIR301~318", "-"],
            ["1.3.3", "5", "변동 후 기업유형", "J7", "Enum", "X", "GIR301~318", "-"],
            ["1.3.3", "6", "소유지분 보유 기업", "K7", "텍스트", "X", "-", "-"],
            ["1.3.3", "7", "변동 전 소유지분(%)", "M7", "숫자(%)", "X", "-", "-"],
            ["1.3.3", "8", "변동 후 소유지분(%)", "P7", "숫자(%)", "X", "-", "-"],
        ],
        "1.4 작성요령": [
            ["", "", "", "", "", "", "", "Control Panel에서 [+] 버튼으로 데이터 행 추가 (4행부터)"],
            ["1.4", "1", "소재지국", "B4", "ISO 국가코드", "X", "KR, JP 등", "-"],
            ["1.4", "2", "하위그룹 유형", "C4", "Enum", "X", "GIR1101~1106", "TypeofSubGroupEnumType"],
            ["1.4", "3", "하위그룹 최상위 TIN", "E4", "텍스트", "X", "-", "-"],
            ["1.4", "4", "과세권 국가", "G4", "ISO 국가코드", "X", "-", "-"],
            ["1.4", "5", "적용면제/제외 사유", "I4", "Enum", "X", "GIR1201~1209", "SafeHarbourEnumType"],
            ["1.4", "6", "실효세율 범위", "K4", "Enum", "X", "GIR1301~1314", "EtrRangeEnumType"],
            ["1.4", "7", "SBIE 추가세액 발생", "M4", "bool", "X", "TRUE/FALSE", "-"],
            ["1.4", "8", "추가세액(QDMTT) 범위", "O4", "Enum", "X", "GIR1401~1409", "QdmtTuTEnumType"],
            ["1.4", "9", "추가세액(GloBE) 범위", "Q4", "Enum", "X", "GIR1501~1509", "GlobeTuTEnumType"],
        ],
    }

    headers = ["섹션", "항목번호", "항목명", "셀 위치", "입력 형식", "복수기재", "허용값 (Enum)", "비고"]
    widths = [10, 10, 25, 12, 18, 10, 50, 50]

    for sheet_name, rows in guides.items():
        ws = wb.create_sheet(sheet_name)
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            ws.column_dimensions[chr(64 + col)].width = w

        for r, row_data in enumerate(rows, 2):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                if c >= 7:
                    cell.alignment = WRAP

        ws.freeze_panes = "A2"


def update_meta_block_counts(wb):
    """_META 시트의 blockCount를 샘플 데이터에 맞게 업데이트"""
    meta = wb["_META"]
    updates = {
        "blockCount:1.3.3": 2,  # 기업구조 변동 2행
        "blockCount:1.4": 3,    # 정보 요약 3행
    }
    row = 2
    while True:
        key = meta.cell(row=row, column=1).value
        if key is None:
            break
        if key in updates:
            meta.cell(row=row, column=2, value=updates[key])
        row += 1


def main():
    shutil.copy2(TEMPLATE, OUTPUT)
    wb = load_workbook(OUTPUT)

    print("샘플 데이터 채움...")
    fill_data(wb)

    print("_META blockCount 업데이트...")
    update_meta_block_counts(wb)

    # 작성요령은 별도 .md 파일로 분리 (시트에 넣지 않음)

    wb.save(OUTPUT)
    print(f"완료: {OUTPUT}")


if __name__ == "__main__":
    main()
