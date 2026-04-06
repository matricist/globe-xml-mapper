"""
템플릿 xlsx 파일 생성 스크립트.
원본 template.xlsx에서 해당 시트를 복사하고 작성요령 시트를 추가.
"""

import os
import shutil
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.join(os.path.dirname(__file__), "..")
SOURCE_FILE = os.path.join(BASE_DIR, "Resources", "template.xlsx")
TEMPLATES_DIR = os.path.join(BASE_DIR, "Resources", "templates")
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")


def copy_sheet_from_source(source_sheet_name, target_wb, target_sheet_name):
    """원본 template.xlsx에서 시트를 복사하여 target_wb에 추가"""
    src_wb = load_workbook(SOURCE_FILE)
    src_ws = src_wb[source_sheet_name]
    tgt_ws = target_wb.active
    tgt_ws.title = target_sheet_name

    # 셀 값, 스타일 복사
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = tgt_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    # 병합 셀 복사
    for merged in src_ws.merged_cells.ranges:
        tgt_ws.merge_cells(str(merged))

    # 열 너비 복사
    for col_letter, dim in src_ws.column_dimensions.items():
        tgt_ws.column_dimensions[col_letter].width = dim.width

    # 행 높이 복사
    for row_num, dim in src_ws.row_dimensions.items():
        tgt_ws.row_dimensions[row_num].height = dim.height

    src_wb.close()


def add_guide_sheet(wb, rows):
    """작성요령 시트 추가 (빈 행 없이)"""
    ws = wb.create_sheet("작성요령")
    headers = ["섹션", "항목번호", "항목명", "셀 위치", "입력 형식", "복수기재", "허용값 (Enum)", "비고"]
    widths = [10, 10, 30, 10, 20, 12, 60, 40]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[chr(64 + col)].width = w

    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 7:
                cell.alignment = WRAP

    ws.freeze_panes = "A2"


def create_filing_info():
    """template_1.1~1.2.xlsx — 원본 국조53부표2 (1) 시트 복사"""
    wb = Workbook()
    copy_sheet_from_source("국조53부표2 (1)", wb, "1.1~1.2")

    guide_rows = [
        ["헤더", "-", "사업연도 시작", "C3", "날짜 (YYYY-MM-DD)", "X", "-", "-"],
        ["헤더", "-", "사업연도 종료", "C6", "날짜 (YYYY-MM-DD)", "X", "-", "-"],
        ["헤더", "-", "법인명(상호)", "P3", "텍스트", "X", "-", "-"],
        ["헤더", "-", "사업자등록번호", "P5", "텍스트", "X", "-", "-"],
        ["1.1", "1", "최종모기업 여부", "B10", "Enum", "X",
         "GIR401=UPE, GIR402=Designated Filing, GIR403=Designated Local, GIR404=CE, GIR405=Other",
         "FilingCeRoleEnumType"],
        ["1.1", "2", "신고구성기업 상호", "D10", "텍스트", "X", "-", "-"],
        ["1.1", "3", "납세자번호", "H10", "텍스트", "X", "-", "-"],
        ["1.1", "4", "신고구성기업 지위", "K10", "Enum", "X",
         "GIR401~GIR405 (상동)", "FilingCeRoleEnumType"],
        ["1.1", "5", "소재지국", "M10", "ISO 국가코드", "X", "KR, JP, US 등", "CountryCodeType"],
        ["1.1", "6", "교환당사국", "O10", "ISO 국가코드", "O (쉼표구분)", "KR, JP, US 등", "복수 시 쉼표로 구분"],
        ["1.2.1", "1", "다국적기업그룹 명칭", "B15", "텍스트", "X", "-", "-"],
        ["1.2.1", "2", "사업연도 개시일", "F15", "날짜", "X", "-", "-"],
        ["1.2.1", "3", "사업연도 종료일", "K15", "날짜", "X", "-", "-"],
        ["1.2.1", "4", "수정신고서 여부", "O15", "Enum", "X",
         "GIR101=신규, GIR102=정정, GIR103=보고대상없음", "MessageTypeIndicEnumType"],
        ["1.2.2", "1", "연결재무제표 유형", "B18", "Enum", "X",
         "GIR501=a, GIR502=b, GIR503=c, GIR504=d", "FilingCeCofUpeEnumType"],
        ["1.2.2", "2", "회계기준", "H18", "텍스트", "X", "-", "예: K-IFRS"],
        ["1.2.2", "3", "통화", "M18", "ISO 통화코드", "X", "KRW, USD 등", "CurrCodeType"],
    ]
    add_guide_sheet(wb, guide_rows)

    path = os.path.join(TEMPLATES_DIR, "template_1.1~1.2.xlsx")
    wb.save(path)
    print(f"  생성: {path}")


def create_upe():
    """template_1.3.1.xlsx — 원본 국조53부표2 (1) 시트 복사 (1.3.1 섹션 포함)"""
    wb = Workbook()
    copy_sheet_from_source("국조53부표2 (1)", wb, "1.3.1")

    guide_rows = [
        ["1.3.1", "1", "소재지국", "J22", "ISO 국가코드", "X", "KR, JP 등", "70010: 하나만"],
        ["1.3.1", "2", "적용가능규칙", "J23", "Enum", "O (쉼표)",
         "GIR201=QIIR(타국), GIR202=QIIR(자국+타국), GIR203=QUTPR, GIR204=QDMTT, GIR205=N/A",
         "IdTypeRulesEnumType"],
        ["1.3.1", "3", "상호", "J24", "텍스트", "X", "-", "-"],
        ["1.3.1", "4", "납세자번호", "J25", "텍스트", "O (쉼표)", "-", "60022: FilingCE TIN 일치 필요"],
        ["1.3.1", "5", "신고접수국가 TIN", "J26", "텍스트", "X", "-", "issuedBy=KR 자동"],
        ["1.3.1", "6", "기업유형", "J27", "Enum", "O (쉼표)",
         "GIR301=CE, GIR302~304=Flow-Through/Hybrid, GIR306=Main Entity, GIR310~311=Investment",
         "70009: UPE에서 GIR305,307-309,312-315,317,318 불가"],
        ["1.3.1", "7", "제외기업 유형", "J28", "Enum", "X",
         "GIR601~606", "ExcludedUpeEnumType"],
        ["1.3.1", "8", "시행령 적용", "J29", "ISO 국가코드", "X", "KR 등", "-"],
    ]
    add_guide_sheet(wb, guide_rows)

    path = os.path.join(TEMPLATES_DIR, "template_1.3.1.xlsx")
    wb.save(path)
    print(f"  생성: {path}")


def create_ce():
    """template_1.3.2.1.xlsx — 원본 국조53부표2 (2) 시트 복사"""
    wb = Workbook()
    copy_sheet_from_source("국조53부표2 (2)", wb, "1.3.2.1")

    guide_rows = [
        ["변동", "1", "변동 여부", "O5", "bool", "X", "TRUE/FALSE, Y/N, 예/아니오", "-"],
        ["소재지국", "2", "소재지국", "O6", "ISO 국가코드", "X", "KR, JP 등", "70011: 하나만"],
        ["소재지국", "3", "적용가능규칙", "O7", "Enum", "O (쉼표)",
         "GIR201~GIR205", "IdTypeRulesEnumType, 70012"],
        ["기본정보", "4", "구성기업 상호", "O8", "텍스트", "X", "-", "-"],
        ["기본정보", "5", "납세자번호", "O9", "텍스트", "X", "-", "-"],
        ["기본정보", "6", "신고접수국가 TIN", "O10", "텍스트", "X", "-", "issuedBy=KR 자동"],
        ["기본정보", "7", "기업유형", "O11", "Enum", "O (쉼표)",
         "GIR301~GIR318", "70013~70021 조합 규칙"],
        ["소유지분", "8", "소유지분 유형", "별첨!B열", "Enum", "X",
         "GIR801=UPE, GIR802=CE, GIR803=JV, GIR804=JV Sub, GIR805=Excluded, GIR806=Non-Group",
         "OwnershipTypeEnumType. 별첨 시트에 복수 행 기재"],
        ["소유지분", "9", "보유기업 TIN", "별첨!C열", "텍스트", "X", "-", "70030. 별첨 시트에 기재"],
        ["소유지분", "10", "소유지분(%)", "별첨!D열", "숫자(%)", "X", "퍼센트로 입력 (예: 80)", "70026~70028. 별첨 시트에 기재"],
        ["QIIR", "11", "모기업 유형", "O16", "Enum", "X",
         "GIR901=POPE, GIR902=IPE, GIR903=Art10.3.5", "PopeipeEnumType"],
        ["QIIR", "12", "QIIR 모기업 TIN", "O17", "텍스트", "X", "-", "70033"],
        ["QIIR", "13", "부분소유모기업 TIN", "O18", "텍스트", "X", "-", "-"],
        ["QUTPR", "14", "해외진출 초기 특례", "O19", "bool", "X", "TRUE/FALSE", "-"],
        ["QUTPR", "15", "소유지분 합계(%)", "O20", "숫자", "X", "0~100 또는 0~1", "-"],
        ["QUTPR", "16", "UPE소유지분>합계", "O21", "bool", "X", "TRUE/FALSE", "-"],
    ]
    add_guide_sheet(wb, guide_rows)

    path = os.path.join(TEMPLATES_DIR, "template_1.3.2.1.xlsx")
    wb.save(path)
    print(f"  생성: {path}")


def create_excluded_entity():
    """template_1.3.2.2.xlsx — 원본 국조53부표2 (3) 시트 복사"""
    wb = Workbook()
    copy_sheet_from_source("국조53부표2 (3)", wb, "1.3.2.2")

    guide_rows = [
        ["제외기업", "1", "변동 여부", "O4", "bool", "X", "TRUE/FALSE, Y/N, 예/아니오", "-"],
        ["제외기업", "2", "제외기업 상호", "O5", "텍스트", "X", "-", "-"],
        ["제외기업", "3", "제외기업 유형", "O6", "Enum", "X",
         "GIR1001=정부기관, GIR1002=국제기구, GIR1003=비영리, GIR1004=연금, GIR1005=투자펀드(UPE), GIR1006=부동산투자(UPE), GIR1007=Art1.5.2(a), GIR1008=Art1.5.2(b)",
         "ExcludedEntityEnumType"],
    ]
    add_guide_sheet(wb, guide_rows)

    path = os.path.join(TEMPLATES_DIR, "template_1.3.2.2.xlsx")
    wb.save(path)
    print(f"  생성: {path}")


if __name__ == "__main__":
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    print("템플릿 생성 중...")
    create_filing_info()
    create_upe()
    create_ce()
    create_excluded_entity()
    print("완료!")
